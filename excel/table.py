import openpyxl as opx
from time import sleep

def strtoArray(str):
    try:
        str = str.replace(" ", "").split(",")
        
        array = []
        
        for n in str:
            array.append(int(n))
            
        return array
    
    except ValueError:
        return False
        

def avisoTabela(data):
    print("Vale lembra que as colunas de dados são: ", end="")
    for i, dado in enumerate(data):
        if i != len(data)-1: 
            print(f"{i} - {dado}, ", end="")
        else:
            print(f"{i} - {dado}")

class ExTable():
    def __init__(self, archive = False):
        if archive == False: 
            self.arquivo = opx.Workbook()
            self.nome = None
        # self.arquivo.remove(self.arquivo.worksheets[0])
        
    def criaTabela(self, tnome):
        self.arquivo.create_sheet(tnome)
        self.arquivo.sheetnames.sort()
        
    def listaTabelas(self):
        return self.arquivo.sheetnames
    
    def tabelaAtiva(self):
        return self.arquivo.active.title
    
    def linhasTabela(self, i):
        tabela = self.arquivo[self.arquivo.sheetnames[i]]
        return tabela.max_row
    
    def colunasTabela(self, i):
        tabela = self.arquivo[self.arquivo.sheetnames[i]]
        return tabela.max_column
        
    
    def addData(self, tabelaIx, columns, data, dataInner = False, condition = False, cab = False):
        
        try: 
            tabela = self.arquivo[self.arquivo.sheetnames[tabelaIx]]
            
            if cab == True:
                cel = tabela.cell(row=(1), column=columns[0])
                cel.value = data
                return True
            
            if condition == False:
                for ix, coluna in enumerate(columns):
                    for dado in range(1, len(data)):
                        cel = tabela.cell(row=(dado+1), column=coluna)
                        cel.value = data[dado][ix]
            
            else:
                for dado in data:
                    procuro = str(dado[condition[0]])
                    i = 2
                    while i <= tabela.max_row:
                        celBusca = tabela.cell(row=i, column = condition[1])
                        # print(celBusca.value , procuro)
                        
                        if ( procuro == str(celBusca.value) ):
                            #if dataInner == False:
                            for nColuna in range(len(columns)):
                                print(dado[dataInner[nColuna]], end=' para = ')
                                cel = tabela.cell(row=i, column=columns[nColuna])
                                print(cel)
                                cel.value = dado[dataInner[nColuna]]
                            i = tabela.max_row
                        
                        i += 1
            return True
        
        except:
            return False 
                    
            
    def saveArchive(self):
        self.arquivo.save(filename=f"{self.nome}.xlsx")
        
    
    def tableItens(self, dataTable):
        print(f"""                    =_=_=_=_=_=_=_= TABELA PESONALIZADA {self.nome} =_=_=_=_=_=_=_=\n""")
        data = dataTable
        while True:
            optext = f"""\n
                    --- Para fazer a manipulacao de arquivos escolha uma opcao abaixo ---
                    Lista de tabelas: """
            if (len(self.arquivo.sheetnames)) != 0:
                for i, tbl in enumerate(self.listaTabelas()):
                    optext += f"""\n                    {i} - {tbl}"""
            else:
                optext += "\nNenhuma tabela ativa"
                
            optext +=f"""
                    \n
                    [0] Voltar para extracao de dados;
                    [1] Mudar nome do arquivo;
                    [2] Criar tabela em {self.nome}.xlsx;
                    [3] Listar areas de trabalho existentes;
                    [4] Adicionar dados a uma tabela simples;
                    [5] Adicionar dados condicionais a uma tabela;
                    [6] Salvar Arquivo ( Possiblita vizualizar o mesmo antes de fechar a edicao ;-) );
                        
                    Rs:"""
            option = int(input(optext))
            
            match option:
                case 0:
                    return
                
                case 1:
                    self.nome = input("\nNome do arquvio: ")
                    
                case 2:
                    self.criaTabela(input("\nNome da Tabela: "))
                    
                case 3:
                    print(self.listaTabelas())
                    
                case 4: 
                    try:
                        tabelaIX = int(input("\nEscolha a tabela a ter dados inseridos: ").replace(" ", ""))
                        
                        avisoTabela(dataTable[0])
                        
                        columns = input("Digite as colunas onde quer inserir na ordem dos dados (',' entre elas): ")
                        columns = strtoArray(columns)
                        
                        self.addData( tabelaIx = tabelaIX, columns = columns, data = data)
                    except:
                        print("Algum dado foi inserido da forma incorreta ou houve um erro desconhecido, refaca o processo")
                        sleep(4)
                        
                    
                case 5:
                    try:
                        tabelaIX = int(input("\nEscolha a tabela a ter dados inseridos: ").replace(" ", ""))
                        
                        avisoTabela(dataTable[0])
                        
                        dataInner = input("Digite o índice dos dados a serem inseridos na tabela (',' entre eles): ")
                        dataInner = strtoArray(dataInner)
                        
                        columns = input("Digite as colunas onde quer inserir na ordem dos dados que quer inserir (',' entre elas): ")
                        columns = strtoArray(columns)
                        
                        conditions = input("['dados extraido indice', 'coluna indice'] analisar dado e coluna (',' entre elas): ")
                        conditions = strtoArray(conditions)
                        
                        if(not ( self.addData( tabelaIx = tabelaIX, columns = columns, data = data, condition = conditions, dataInner = dataInner) ) ):
                            print("Error: Houve algum erro na geracao da tabela, refaca o processo de insercao")
                    except:
                        print("Algum dado foi inserido da forma incorreta, refaca o processo")
                        sleep(4)
            
                case 6:
                    self.saveArchive()
